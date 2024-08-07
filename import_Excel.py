import os
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import quote_sheetname
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule

import globalvars

class field_notes():
    def __init__(self) -> None:
        pass

    def get_notes_path(self, dictionary):
        # define the main directory path
        wb_dir = globalvars.root_path + '/' + dictionary['Structure ID'] + '/Field Notes'
        # define an empty path for workbook name
        wb_name = ''
        # iterate over all files and folders in the directory
        for file in os.listdir(wb_dir):
            # if the file is an excel, the get the name of it
            if '.xlsx' in file:
                wb_name = file
                break
        # construct the path of the workbook
        wb_path = wb_dir + '/' + wb_name
        return wb_path

    def import_notes(self, dictionary):
        # get the workbook path
        wb_path = self.get_notes_path(dictionary)
        # load the workbook
        wb = openpyxl.load_workbook(wb_path)
        # get all sheets in the workbook
        wb_sheetnames = wb.sheetnames
        # get all elements
        wb_elements = wb_sheetnames[1:(len(wb_sheetnames) - 2)]
        # update the dictionary with the elements
        secondary_elements = []
        for i in range(len(wb_elements)):
            if wb_elements[i] in ['811', '812', '813', '814', '815', '816', '510', '515', '520', '521']:
                secondary_elements.append(wb_elements[i])
            else:
                main_element = {wb_elements[i]: secondary_elements}
                dictionary['Elements'].append(main_element)
                secondary_elements = []

        #print(dictionary['Structure ID'])
        #print(dictionary['Elements'])
        #
